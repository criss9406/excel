"""
Generador de reportes Excel a partir del resultado de un analisis.

Para EDA, el archivo de salida tiene 5 hojas:
  - Resumen, Perfil columnas, Pasos del proceso, Hallazgos, Datos.

Para Inventario, las hojas son:
  - Resumen, Parametros, Riesgo de quiebre, Sobre-stock, ABC, Productos zombi,
    Pasos del proceso, Datos.

Se entrega como bytes en memoria, listo para StreamingResponse.
"""
from __future__ import annotations
from io import BytesIO
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
_CELL_ALIGN = Alignment(vertical="top", wrap_text=False)

# Topes razonables para el auto-ancho (en unidades de Excel ~ caracteres)
_ANCHO_MIN = 12
_ANCHO_MAX = 55
_MUESTRA_FILAS = 200  # cuantas filas mirar para calcular ancho


def _estilizar_hoja(ws, df: pd.DataFrame, con_filtro: bool) -> None:
    """Aplica encabezado destacado, freeze, anchos y filtro a una hoja."""
    if ws.max_row == 0 or df.empty:
        return

    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    for idx, col in enumerate(df.columns, start=1):
        muestra = df[col].head(_MUESTRA_FILAS).fillna("").astype(str)
        max_len = max((len(v) for v in muestra), default=0)
        ancho = min(max(max_len, len(str(col))) + 2, _ANCHO_MAX)
        ws.column_dimensions[get_column_letter(idx)].width = max(ancho, _ANCHO_MIN)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _CELL_ALIGN

    if con_filtro and ws.max_row > 1:
        ultima_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{ultima_col}{ws.max_row}"


def _hoja_resumen(resultado: dict[str, Any]) -> pd.DataFrame:
    filas = [
        ("Archivo", resultado.get("archivo", "")),
        ("Filas", resultado.get("n_filas", 0)),
        ("Columnas", resultado.get("n_columnas", 0)),
    ]
    if resultado.get("tipo_analisis") == "inventario":
        inv = resultado.get("kpis_inventario", {})
        filas.extend([
            ("Productos en riesgo de quiebre", inv.get("riesgo_quiebre", {}).get("n_productos", 0)),
            ("Valor en riesgo (estimado)", inv.get("riesgo_quiebre", {}).get("valor_total_en_riesgo", 0)),
            ("Productos con sobre-stock", inv.get("sobre_stock", {}).get("n_productos", 0)),
            ("Capital inmovilizado por sobre-stock", inv.get("sobre_stock", {}).get("capital_inmovilizado_total", 0)),
            ("Productos zombi", inv.get("zombis", {}).get("n_productos", 0)),
            ("Capital inmovilizado en zombis", inv.get("zombis", {}).get("valor_inmovilizado_total", 0)),
            ("Ingreso total (universo)", inv.get("abc", {}).get("ingreso_total", 0)),
        ])
    else:
        filas.extend([
            ("Total violaciones detectadas", resultado["resumen"]["total_violaciones"]),
            ("Columnas con violaciones", resultado["resumen"]["columnas_con_violaciones"]),
            ("Pasos en severidad info", resultado["resumen"]["por_severidad"].get("info", 0)),
            ("Pasos en severidad advertencia", resultado["resumen"]["por_severidad"].get("advertencia", 0)),
            ("Pasos en severidad error", resultado["resumen"]["por_severidad"].get("error", 0)),
        ])
    return pd.DataFrame(filas, columns=["Metrica", "Valor"])


def _hoja_parametros_inv(resultado: dict[str, Any]) -> pd.DataFrame:
    p = resultado.get("parametros", {})
    m = resultado.get("mapeo_columnas", {})
    filas = [
        ("Lead time (dias)", p.get("lead_time_dias")),
        ("Margen sobre precio (%)", p.get("margen_pct")),
        ("Ventana zombi (dias)", p.get("dias_zombi")),
        ("Factor sobre-stock (x lead time)", p.get("factor_sobrestock")),
        ("—", "—"),
        ("Columna fecha", m.get("fecha", "(no detectada)")),
        ("Columna producto", m.get("producto", "(no detectada)")),
        ("Columna stock", m.get("stock", "(no detectada)")),
        ("Columna unidades vendidas", m.get("vendidas", "(no detectada)")),
        ("Columna precio", m.get("precio", "(no detectada)")),
    ]
    return pd.DataFrame(filas, columns=["Parametro", "Valor"])


def _hoja_riesgo(resultado: dict[str, Any]) -> pd.DataFrame:
    top = resultado["kpis_inventario"]["riesgo_quiebre"]["top"]
    if not top:
        return pd.DataFrame([{"Mensaje": "Sin productos en riesgo de quiebre"}])
    df = pd.DataFrame(top)
    return df.rename(columns={
        "producto": "Producto",
        "stock_actual": "Stock actual",
        "venta_diaria_prom": "Venta diaria promedio",
        "cobertura_dias": "Cobertura (dias)",
        "dias_faltantes": "Dias faltantes",
        "valor_riesgo": "Valor en riesgo",
    })


def _hoja_sobrestock(resultado: dict[str, Any]) -> pd.DataFrame:
    top = resultado["kpis_inventario"]["sobre_stock"]["top"]
    if not top:
        return pd.DataFrame([{"Mensaje": "Sin productos con sobre-stock"}])
    df = pd.DataFrame(top)
    return df.rename(columns={
        "producto": "Producto",
        "stock_actual": "Stock actual",
        "venta_diaria_prom": "Venta diaria promedio",
        "cobertura_dias": "Cobertura (dias)",
        "dias_excedentes": "Dias excedentes",
        "capital_inmovilizado": "Capital inmovilizado",
    })


def _hoja_abc(resultado: dict[str, Any]) -> pd.DataFrame:
    abc = resultado["kpis_inventario"]["abc"]
    ranking = abc.get("ranking_top", [])
    if not ranking:
        return pd.DataFrame([{"Mensaje": "Sin productos con ingreso registrado"}])
    df = pd.DataFrame(ranking)
    return df.rename(columns={
        "producto": "Producto",
        "ingreso_total": "Ingreso",
        "pct_acumulado": "% acumulado",
        "clase_abc": "Clase",
    })


def _hoja_zombis(resultado: dict[str, Any]) -> pd.DataFrame:
    top = resultado["kpis_inventario"]["zombis"]["top"]
    if not top:
        return pd.DataFrame([{"Mensaje": "Sin productos zombi"}])
    df = pd.DataFrame(top)
    return df.rename(columns={
        "producto": "Producto",
        "stock_actual": "Stock actual",
        "ultima_venta": "Ultima venta",
        "dias_sin_venta": "Dias sin venta",
        "valor_inmovilizado": "Capital inmovilizado",
    })


def _hoja_perfil(resultado: dict[str, Any]) -> pd.DataFrame:
    perfil = resultado.get("perfil_columnas", [])
    if not perfil:
        return pd.DataFrame([{"Mensaje": "Sin perfil"}])
    df = pd.DataFrame(perfil)
    df["ejemplos"] = df["ejemplos"].apply(lambda lst: ", ".join(lst) if isinstance(lst, list) else "")
    return df


def _hoja_pasos(resultado: dict[str, Any]) -> pd.DataFrame:
    pasos = resultado.get("traza", [])
    filas = []
    for p in pasos:
        filas.append({
            "Paso": p["nombre"],
            "Columna": p["columna"] or "(varias)",
            "Regla": p["regla"],
            "Parametros": ", ".join(f"{k}={v}" for k, v in p["params"].items()),
            "Alcance": p["alcance"],
            "Filas en alcance": p["n_alcance"],
            "Violaciones": p["n_violaciones"],
            "Severidad": p["severidad"],
        })
    return pd.DataFrame(filas) if filas else pd.DataFrame([{"Mensaje": "Sin pasos"}])


def _hoja_hallazgos(resultado: dict[str, Any]) -> pd.DataFrame:
    """Expande cada paso con violaciones en filas individuales (una por ejemplo)."""
    filas = []
    for p in resultado.get("traza", []):
        if p["n_violaciones"] == 0:
            continue
        for idx, ejemplo in zip(p["indices"][:50], p["ejemplos"]):
            valor = ejemplo.get(p["columna"]) if p["columna"] else None
            filas.append({
                "Fila (indice)": idx,
                "Columna": p["columna"] or "(fila completa)",
                "Regla": p["regla"],
                "Valor": valor,
                "Paso": p["nombre"],
            })
    return pd.DataFrame(filas) if filas else pd.DataFrame([{"Mensaje": "Sin hallazgos"}])


def _hoja_datos_marcados(resultado: dict[str, Any]) -> pd.DataFrame:
    """Datos originales con columna `Errores detectados` listando reglas por fila."""
    df_dict = resultado.get("df_dict", [])
    if not df_dict:
        return pd.DataFrame([{"Mensaje": "Sin datos"}])
    df = pd.DataFrame(df_dict)

    reglas_por_indice: dict[int, list[str]] = {}
    for p in resultado.get("traza", []):
        if p["n_violaciones"] == 0:
            continue
        for idx in p["indices"]:
            reglas_por_indice.setdefault(idx, []).append(p["nombre"])

    df["Errores detectados"] = [
        " | ".join(reglas_por_indice.get(i, [])) for i in range(len(df))
    ]
    return df


def _hoja_cambios(resultado: dict[str, Any]) -> pd.DataFrame:
    cambios = resultado.get("cambios_limpieza", [])
    if not cambios:
        return pd.DataFrame([{"Mensaje": "Sin cambios"}])
    return pd.DataFrame(cambios)

def generar_excel(resultado: dict[str, Any]) -> bytes:
    """Construye el archivo Excel completo y devuelve sus bytes."""
    if resultado.get("tipo_analisis") == "inventario":
        df_dict = resultado.get("df_dict", [])
        df_limpio = pd.DataFrame(df_dict) if df_dict else pd.DataFrame([{"Mensaje": "Sin datos"}])
        
        hojas = [
            ("Datos Limpios", df_limpio, True),
            ("Cambios", _hoja_cambios(resultado), True),
        ]
    else:
        hojas = [
            ("Resumen", _hoja_resumen(resultado), False),
            ("Perfil columnas", _hoja_perfil(resultado), True),
            ("Pasos del proceso", _hoja_pasos(resultado), True),
            ("Hallazgos", _hoja_hallazgos(resultado), True),
            ("Datos", _hoja_datos_marcados(resultado), True),
        ]
    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        for nombre, df, con_filtro in hojas:
            df.to_excel(writer, sheet_name=nombre, index=False)
            _estilizar_hoja(writer.sheets[nombre], df, con_filtro=con_filtro)
    salida.seek(0)
    return salida.read()

def generar_pdf_inventario(resultado: dict[str, Any]) -> bytes:
    """Genera un reporte PDF con los dashboards del inventario."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elementos = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    elementos.append(Paragraph(f"Dashboard de Inventario: {resultado.get('archivo', '')}", title_style))
    elementos.append(Spacer(1, 12))
    
    # Resumen
    elementos.append(Paragraph("Resumen General", subtitle_style))
    inv = resultado.get("kpis_inventario", {})
    glob = inv.get("globales", {})
    datos_resumen = [
        ["Total Productos (Filas)", str(resultado.get("n_filas", 0))],
        ["Columnas", str(resultado.get("n_columnas", 0))],
        ["Ingreso Total ABC", f"${inv.get('abc', {}).get('ingreso_total', 0):,.0f}".replace(',', '.')],
        ["Rotación (Veces)", str(glob.get("rotacion_unidades", "-"))],
        ["Cobertura Global (Días)", str(glob.get("cobertura_global_dias", "-"))],
    ]
    t_resumen = Table(datos_resumen, colWidths=[200, 200])
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(t_resumen)
    elementos.append(Spacer(1, 12))
    
    # Salud del Inventario
    elementos.append(Paragraph("Distribución de Salud del Inventario", subtitle_style))
    salud = inv.get("salud", {})
    if salud:
        datos_salud = [
            ["Estado", "N° Productos", "Capital Inmovilizado ($)"],
            ["Saludable", str(salud.get("saludable", {}).get("n", 0)), f"{salud.get('saludable', {}).get('capital', 0):,.0f}".replace(',', '.')],
            ["Riesgo Quiebre", str(salud.get("riesgo_quiebre", {}).get("n", 0)), f"{salud.get('riesgo_quiebre', {}).get('capital', 0):,.0f}".replace(',', '.')],
            ["Sobre-stock", str(salud.get("sobre_stock", {}).get("n", 0)), f"{salud.get('sobre_stock', {}).get('capital', 0):,.0f}".replace(',', '.')],
            ["Zombis", str(salud.get("zombi", {}).get("n", 0)), f"{salud.get('zombi', {}).get('capital', 0):,.0f}".replace(',', '.')]
        ]
        t_salud = Table(datos_salud)
        t_salud.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elementos.append(t_salud)
    elementos.append(Spacer(1, 12))
    
    # Riesgo de Quiebre
    elementos.append(Paragraph("1. Riesgo de Quiebre (Top 10)", subtitle_style))
    riesgo_top = inv.get("riesgo_quiebre", {}).get("top", [])
    if riesgo_top:
        datos_riesgo = [["Producto", "Stock", "Venta Diaria", "Dias Faltantes", "Valor Riesgo"]]
        for r in riesgo_top[:10]:
            datos_riesgo.append([str(r["producto"]), str(r["stock_actual"]), str(r["venta_diaria_prom"]), str(r["dias_faltantes"]), str(r["valor_riesgo"])])
        t_riesgo = Table(datos_riesgo)
        t_riesgo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elementos.append(t_riesgo)
    else:
        elementos.append(Paragraph("Sin productos en riesgo.", normal_style))
    elementos.append(Spacer(1, 12))
    
    # Sobre-stock
    elementos.append(Paragraph("2. Sobre-stock (Top 10)", subtitle_style))
    sobre_top = inv.get("sobre_stock", {}).get("top", [])
    if sobre_top:
        datos_sobre = [["Producto", "Stock", "Venta Diaria", "Dias Excedentes", "Capital Inmovilizado"]]
        for r in sobre_top[:10]:
            datos_sobre.append([str(r["producto"]), str(r["stock_actual"]), str(r["venta_diaria_prom"]), str(r["dias_excedentes"]), str(r["capital_inmovilizado"])])
        t_sobre = Table(datos_sobre)
        t_sobre.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elementos.append(t_sobre)
    else:
        elementos.append(Paragraph("Sin productos con sobre-stock.", normal_style))
    elementos.append(Spacer(1, 12))
    
    # Productos Zombi
    elementos.append(Paragraph("3. Productos Zombi (Top 10)", subtitle_style))
    zombi_top = inv.get("zombis", {}).get("top", [])
    if zombi_top:
        datos_zombi = [["Producto", "Stock", "Dias sin Venta", "Capital Inmovilizado"]]
        for r in zombi_top[:10]:
            datos_zombi.append([str(r["producto"]), str(r["stock_actual"]), str(r.get("dias_sin_venta", "-")), str(r["valor_inmovilizado"])])
        t_zombi = Table(datos_zombi)
        t_zombi.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elementos.append(t_zombi)
    else:
        elementos.append(Paragraph("Sin productos zombi.", normal_style))

    doc.build(elementos)
    buffer.seek(0)
    return buffer.read()
